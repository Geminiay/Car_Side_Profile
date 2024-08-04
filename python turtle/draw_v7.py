import turtle
import random
import math
from sympy import symbols, solve
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import os
from datetime import datetime

#Constants
totalWidth = 4147.65
totalHeight = 1540.87
base = 4147.65
roof = 2493.97
backHeight = 850.54
hood = 1150.29
hoodAngle = 4.68
frontHeight = 871.27
hoodCos = math.cos(math.radians(hoodAngle))
hoodSin = math.sin(math.radians(hoodAngle))
workbook = openpyxl.Workbook()
sheet = workbook.active
current_date = datetime.now().strftime('%Y_%m_%d')
#First line is headers and the second line is the original parameters
next_row = 2

def calculateParameters():
    #Parameters
    x, y, z = symbols('x y z')
    backAngle = round(random.uniform(70,85), 2) #actual value is 77.73
    frontAngle = round(random.uniform(50,65), 2) #actual value is 58.19
    frontCos = math.cos(math.radians(frontAngle))
    frontSin = math.sin(math.radians(frontAngle))
    backCos = math.cos(math.radians(backAngle))
    backSin = math.sin(math.radians(backAngle))
    
    #Calculations
    widthEq = hood*hoodCos+x*frontCos+z+y*backCos-base
    heightEq1 = frontHeight+hood*hoodSin+x*frontSin-totalHeight
    heightEq2 = backHeight+y*backSin-totalHeight
    eq = solve((widthEq, heightEq1, heightEq2),(x, y, z),dict=True)
    windShield = round(eq[0][x],2) #actual value is 677.49
    rearWindow = round(eq[0][y],2) #actual value is 706.46
    roof = round(eq[0][z],2) #actual value is 2493.97
    
    return windShield, frontAngle, rearWindow, backAngle, roof

def draw(windShield, frontAngle, rearWindow, backAngle, roof):
    
    #Disables the function if withImage is false
    if withImage == False:
        return
    
    turtle.title(f"Vehicle Side Profile Number {next_row-2}")
    
    #Draw
    t.hideturtle()
    t.penup()
    t.goto(-totalWidth/10, -totalHeight/10)
    t.pendown()
    t.fd(base/5)
    t.lt(90)
    t.fd(backHeight/5)
    t.lt(90-backAngle)
    t.fd(rearWindow/5)
    t.lt(backAngle)
    t.fd(roof/5)
    t.lt(frontAngle)
    t.fd(windShield/5)
    t.rt(frontAngle-hoodAngle)
    t.fd(hood/5)
    t.lt(90-hoodAngle)
    t.fd(frontHeight/5)
    t.lt(90)
    
    #Generate a unique filename based on the current row number
    eps_filename = f'{epsFile}/drawing_{next_row-2}.eps'
    png_filename = f'{imageFile}/drawing_{next_row-2}.png'
    
    #Save the drawing as an EPS file and convert EPS file to PNG
    canvas.postscript(file=eps_filename)
    with Image.open(eps_filename) as img:
        img.save(png_filename)
    t.clear()

def saveResults(windShield, frontAngle, rearWindow, backAngle, roof):
    
    global next_row
    #Print Parameter Results 
    print("Wind Shield Length: ",windShield)
    print("Wind Shield Angle: ",frontAngle)
    print("Rear Window Length: ",rearWindow)
    print("Rear Window Angle: ",backAngle)
    print("Roof Length: ",roof)

    #Insert the image and parameters into the Excel file
    sheet.cell(row=next_row, column=1).value = next_row-2
    sheet.cell(row=next_row, column=2).value = rearWindow
    sheet.cell(row=next_row, column=3).value = windShield
    sheet.cell(row=next_row, column=4).value = roof
    sheet.cell(row=next_row, column=5).value = backAngle
    sheet.cell(row=next_row, column=6).value = frontAngle
    
    #Center-align all text in the newly added row
    for col in range(1, 7):
        sheet.cell(row=next_row, column=col).alignment = Alignment(horizontal='center')

    workbook.save(datasetfile)
    workbook.close()

    print(f"Outputs saved to Excel as parameter number {next_row-2}.")
    next_row += 1

#Getting input for iteration from user
iteration = int(input("How many iterations needed?:\n"))
withImage = input("Do you want images? (y/n):\n").lower().strip() == 'y'

#Create files
parentFile = f'dataset-{current_date}'
datasetfile = f'{parentFile}/dataset.xlsx'
os.makedirs(parentFile)
if withImage == True:
    #Create eps and image files
    epsFile = f'{parentFile}/eps_images'
    imageFile = f'{parentFile}/images'
    os.makedirs(imageFile)
    os.makedirs(epsFile)
    
    #Create turtle screen
    s = turtle.getscreen()
    t = turtle.Turtle()
    screen = turtle.Screen()
    canvas = turtle.getcanvas()
    screen.setup(width=920, height=400)
    turtle.speed(10)

#Ordering the header row
sheet.append(["Parameter Order", "Rear Window Length", "Wind Shield Length", "Roof Length", "Rear Window Angle (70-85)", "Wind Shield Angle (50-65)", "Cd" ])
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
for idx, col in enumerate(columns, start=1):
    sheet.column_dimensions[col].width = 27
    sheet.cell(row=1, column=idx).alignment = Alignment(horizontal='center')

#Draws the original values and saves    
draw(677.49, 58.19, 706.46, 77.73, 2493.97)
saveResults(677.49, 58.19, 706.46, 77.73, 2493.97)  
    
#Iteration
x = 0
while x < iteration:
    windShield, frontAngle, rearWindow, backAngle, roof = calculateParameters()   
    draw(windShield, frontAngle, rearWindow, backAngle, roof)
    saveResults(windShield, frontAngle, rearWindow, backAngle, roof)
    x += 1

print(f"Dataset has been created into {parentFile} file.")