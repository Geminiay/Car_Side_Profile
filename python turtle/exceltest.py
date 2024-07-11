import openpyxl
from random import randint

# Function to generate random outputs
def generate_outputs():
    output1 = randint(1, 100)
    output2 = randint(1, 100)
    output3 = randint(1, 100)
    output4 = randint(1, 100)
    return output1, output2, output3, output4

# Function to save outputs to Excel
def save_to_excel(outputs):
    try:
        workbook = openpyxl.load_workbook("outputs.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active
    next_row = sheet.max_row + 1

    # Write outputs to Excel
    sheet.cell(row=next_row, column=1).value = outputs[0]
    sheet.cell(row=next_row, column=2).value = outputs[1]
    sheet.cell(row=next_row, column=3).value = outputs[2]
    sheet.cell(row=next_row, column=4).value = outputs[3]

    workbook.save("outputs.xlsx")

# Generate outputs
outputs = generate_outputs()

# Save outputs to Excel
save_to_excel(outputs)

print("Outputs saved to Excel.")
