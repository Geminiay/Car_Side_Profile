import turtle
import random
import math
from sympy import symbols, solve
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import os

# Create the 'images' and 'eps_images' directory if it doesn't exist
if not os.path.exists('images'):
    os.makedirs('images')
if not os.path.exists('eps_images'):
    os.makedirs('eps_images')

#Variables
turtle.title("Vehicle Side Profile")
screen = turtle.Screen()
s = turtle.getscreen()
t = turtle.Turtle()
canvas = turtle.getcanvas()
workbook = openpyxl.Workbook()
sheet = workbook.active

x, y, z = symbols('x y z')

workbook = openpyxl.load_workbook("Results.xlsx")
next_row = sheet.max_row + 1


# Generate a unique filename based on the current row number
eps_filename = f'eps_images/drawing_{next_row}.eps'
png_filename = f'images/drawing_{next_row}.png'


#Constants
totalWidth = 4166.03
totalHeight = 1535.63
base = 4166.03
roof = 2507.93
backHeight = 833.01
hood = 1158.16
hoodAngle = 4.88
frontHeight = 862.07


#Parameters
backAngle = round(random.uniform(70,85), 2) #actual value is 78.26
frontAngle = round(random.uniform(50,65), 2) #actual value is 58.19


#Angles
hoodCos = math.cos(math.radians(hoodAngle))
hoodSin = math.sin(math.radians(hoodAngle))

frontCos = math.cos(math.radians(frontAngle))
frontSin = math.sin(math.radians(frontAngle))
backCos = math.cos(math.radians(backAngle))
backSin = math.sin(math.radians(backAngle))


#Calculations
widthEq = hood*hoodCos+x*frontCos+z+y*backCos-base
heightEq1 = frontHeight+hood*hoodSin+x*frontSin-totalHeight
heightEq2 = backHeight+y*backSin-totalHeight
eq = solve((widthEq, heightEq1, heightEq2),(x, y, z),dict=True)
windShield = round(eq[0][x],2) #actual value is 664.75
rearWindow = round(eq[0][y],2) #actual value is 755.63
roof = round(eq[0][z],2) #actual value is 2507.93


#Print Parameter Results 
print("Wind Shield Length: ",windShield)
print("Wind Shield Angle: ",frontAngle)
print("Rear Window Length: ",rearWindow)
print("Rear Window Angle: ",backAngle)
print("Roof Length: ",roof)


# Create turtle screen
screen.setup(width=460, height=200)


#Draw
t.hideturtle()
t.penup()
t.goto(-totalWidth/20, -totalHeight/20)
t.pendown()
t.fd(base/10)
t.lt(90)
t.fd(backHeight/10)
t.lt(90-backAngle)
t.fd(rearWindow/10)
t.lt(backAngle)
t.fd(roof/10)
t.lt(frontAngle)
t.fd(windShield/10)
t.rt(frontAngle-hoodAngle)
t.fd(hood/10)
t.lt(90-hoodAngle)
t.fd(frontHeight/10)


# Save the drawing as an EPS file and convert EPS file to PNG
canvas.postscript(file=eps_filename)
screen.bye()
with Image.open(eps_filename) as img:
    img.save(png_filename)

# Insert the image and parameters into the Excel file
sheet.cell(row=next_row, column=1).value = "{:.2f}".format(windShield)
sheet.cell(row=next_row, column=2).value = "{:.2f}".format(frontAngle)
sheet.cell(row=next_row, column=3).value = "{:.2f}".format(rearWindow)
sheet.cell(row=next_row, column=4).value = "{:.2f}".format(backAngle)
sheet.cell(row=next_row, column=5).value = "{:.2f}".format(roof)
sheet.cell(row=next_row, column=7).value = png_filename

workbook.save("Results.xlsx")
workbook.close()

print(f"Outputs saved to Excel in row {next_row}.")