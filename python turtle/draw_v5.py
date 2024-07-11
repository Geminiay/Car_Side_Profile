import turtle
import random
import math
from sympy import symbols, solve
import openpyxl


turtle.title("Vehicle Side Profile")
s = turtle.getscreen()
t = turtle.Turtle()
x, y, z = symbols('x y z')
workbook = openpyxl.load_workbook("Results.xlsx")
sheet = workbook.active
next_row = sheet.max_row + 1


#Constants
totalWidth = 416.603
totalHeight = 153.563
base = 416.603
roof = 250.793
backHeight = 83.301
hood = 115.816
hoodAngle = 4.88
frontHeight = 86.207


#Parameters
backAngle = round(random.uniform(70,85), 2) #actual value is 78.26
frontAngle = round(random.uniform(50,65), 2) #actual value is 58.19


#Angles
hoodCos = math.cos(math.radians(hoodAngle))
hoodSin = math.sin(math.radians(hoodAngle))
frontCos = math.cos(math.radians(frontAngle))
frontSin = math.sin(math.radians(frontAngle))
backCos = math.cos(math.radians(backAngle))
backSin = math.sin(math.radians(backAngle))


#Calculations
widthEq = hood*hoodCos+x*frontCos+z+y*backCos-base
heightEq1 = frontHeight+hood*hoodSin+x*frontSin-totalHeight
heightEq2 = backHeight+y*backSin-totalHeight
eq = solve((widthEq, heightEq1, heightEq2),(x, y, z),dict=True)
windShield = round(eq[0][x],3) #actual value is 66.475
rearWindow = round(eq[0][y],3) #actual value is 75.563
roof = round(eq[0][z],3) #actual value is 250.793

#Print Parameter Results 
print("Wind Shield Length: ",windShield)
print("Wind Shield Angle: ",frontAngle)
print("Rear Window Length: ",rearWindow)
print("Rear Window Angle: ",backAngle)
print("Roof Length: ",roof)


#Write Outputs to Excel
sheet.cell(row=next_row, column=1).value = "{:.3f}".format(windShield)
sheet.cell(row=next_row, column=2).value = "{:.2f}".format(frontAngle)
sheet.cell(row=next_row, column=3).value = "{:.3f}".format(rearWindow)
sheet.cell(row=next_row, column=4).value = "{:.2f}".format(backAngle)
sheet.cell(row=next_row, column=5).value = "{:.3f}".format(roof)
workbook.save("Results.xlsx")
print("Outputs saved to Excel.")


#Pre-Draw
t.penup()
t.lt(90)
t.fd(50)
t.lt(90)
t.fd(260)
t.lt(180)
t.pendown()


#Draw
t.fd(base)
t.lt(90)
t.fd(backHeight)
t.lt(90-backAngle)
t.fd(rearWindow)
t.lt(backAngle)
t.fd(roof)
t.lt(frontAngle)
t.fd(windShield)
t.rt(frontAngle-hoodAngle)
t.fd(hood)
t.lt(90-hoodAngle)
t.fd(frontHeight)
t.hideturtle()
turtle.done()